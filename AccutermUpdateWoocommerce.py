from woocommerce import API
import csv
import sqlite3
#from openpyxl import load_workbook

import traceback, sys
from PyQt5 import QtWidgets
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.uic import loadUi

from random import randint
from openpyxl import load_workbook

#=================================================================================

class WorkerSignals(QObject):
    '''
    Defines the signals available from a running worker thread.

    Supported signals are:

    finished
        No data

    error
        tuple (exctype, value, traceback.format_exc() )

    result
        object data returned from processing, anything

    progress
        int indicating % progress

    '''
    finished = pyqtSignal()
    error = pyqtSignal(tuple)
    result = pyqtSignal(object)
    progress = pyqtSignal(int)


class Worker(QRunnable):
    '''
    Worker thread

    Inherits from QRunnable to handler worker thread setup, signals and wrap-up.

    :param callback: The function callback to run on this worker thread. Supplied args and
                     kwargs will be passed through to the runner.
    :type callback: function
    :param args: Arguments to pass to the callback function
    :param kwargs: Keywords to pass to the callback function

    '''

    def __init__(self, fn, *args, **kwargs):
        super(Worker, self).__init__()

        # Store constructor arguments (re-used for processing)
        self.fn = fn
        self.args = args
        self.kwargs = kwargs
        self.signals = WorkerSignals()

        # Add the callback to our kwargs
        self.kwargs['progress_callback'] = self.signals.progress

    @pyqtSlot()
    def run(self):
        '''
        Initialise the runner function with passed args, kwargs.
        '''

        # Retrieve args/kwargs here; and fire processing using them
        try:
            result = self.fn(*self.args, **self.kwargs)
        except:
            traceback.print_exc()
            exctype, value = sys.exc_info()[:2]
            self.signals.error.emit((exctype, value, traceback.format_exc()))
        else:
            self.signals.result.emit(result)  # Return the result of the processing
        finally:
            self.signals.finished.emit()  # Done

class SQL():

    def __init__(self, database_name, drop_tables=True):
      try:

        #Connect or Create database_name.db
        self.conn = sqlite3.connect(f'{database_name}.db')
        self.cursor = self.conn.cursor()

        #Drop all existing tables
        if drop_tables: self._drop_tables()

      except sqlite3.Error as err:
        print("Something went wrong: {}".format(err))
      
    #=========================== HELPER FUNCTIONS =============================================

    #Prints sqlite3.connect error
    def _print_error(self, error):
        print("Error:", error)
        print("Error Args:", error.args)

    #Drop all existing tables
    def _drop_tables(self):
        self.cursor.execute('''
        SELECT name FROM sqlite_master 
        WHERE type='table';
        ''')
        for table in self.cursor.fetchall(): 
          try:
            print(f"Deleting `{table[0]}` ...")
            self.cursor.execute(f"DROP TABLE {table[0]};")
          except sqlite3.Error as error:  self._print_error(error)

    #Returns SQL-Formatted CREATE TABLE statement as a string key value
    def _table_constructor(self, table_name): 
        TABLES = {}

        TABLES = dict.fromkeys(['valid_items', 'raw_items'], f"""
        CREATE TABLE {table_name} (
          product_id VARCHAR(11) NOT NULL PRIMARY KEY,
          product_description VARCHAR(100),
          description2 VARCHAR(100),
          description3 VARCHAR(100),
          reg_price INT(11) NOT NULL,
          pu_price INT(11) NOT NULL,
          avail INT(11) NOT NULL,
          is_continuing BOOLEAN NOT NULL
          )""")

        TABLES['website_items'] = ("""
          CREATE TABLE website_items (
            id INT NOT NULL PRIMARY KEY,
            name VARCHAR(11) NOT NULL,
            in_stock BOOLEAN NOT NULL,
            reg_price INT(11) DEFAULT 0
          )""" )
        

        return TABLES[table_name]

    #Return list of column names in 'table_name
    def _get_table_columns(self, table_name):
        try:
          self.cursor.execute(f"SELECT * FROM {table_name}")
          field_names = [i[0] for i in self.cursor.description]
          return field_names

        except sqlite3.Error as error:
          self._print_error(error)

    #Creates sqlite3 'table_name' from a pre-written SQL-formatted dictionary in '_table_constructor'
    def _create_table(self, table_name):
        try:
          self.cursor.execute(f"{self._table_constructor(table_name)};")

        except sqlite3.Error as error:
          self._print_error(error)

    #Inserts 'values' into any existing 'table_name'
    def _insert_table(self, table_name, values):
        try:

          #Stores list of column names into string from table_name
          table_columns = f"{self._get_table_columns(table_name)}"
          text_table = table_columns.maketrans('', '', '[]\'')
          table_columns = table_columns.translate(text_table)

          #Insert Data into Table
          self.cursor.execute(f"""
            INSERT INTO {table_name}
            ({table_columns})
            VALUES {values}
            """)

        except sqlite3.Error as error:  self._print_error(error)

    #Returns SQL Select statement with conditions to filter out invalid items
    def _sql_get_valid_items(self, table_name):
        return f'''
        SELECT *
        FROM {table_name}
        WHERE product_id NOT LIKE '%-SR'
          AND product_id Not Like '%-DRM'
          AND product_id Not Like '%-DR'
          AND product_id Not Like '%-C'
          AND product_id Not Like '%-M'
          AND product_id Not Like '%-N'
          AND product_id Not Like '%-NS'
          AND product_id Not Like '%-NS'
          AND description2 Not Like '% OF %'
          AND description3 Not Like '% OF %'
          AND description3 Not Like '%REPLACEMENT%'
          AND product_id Not Like '%-1'
          AND product_id Not Like '%-2'
          AND product_id Not Like '%-3'

        UNION

        SELECT *
        FROM {table_name}
        WHERE product_description Like '%SET%'
            OR product_description Like '%SINGLE%'
            OR product_description Like '% BED %'
        '''

    #Return SQL Select statement CASE,WHEN,END AS to create new fields 'In stock?' and 'Regular price'
    def _sql_get_stock_price(self, table_name):
        return f'''
        SELECT {table_name}.product_id, 
          CASE 
            WHEN {table_name}.avail > 0 AND ({table_name}.reg_price > 2 OR {table_name}.pu_price > 2) THEN 1
            ELSE 0
          END AS `In stock?`,
          CASE 
            WHEN {table_name}.reg_price > {table_name}.pu_price THEN {table_name}.reg_price
            ELSE {table_name}.pu_price
          END AS `Regular price`
        FROM {table_name} 
        '''

    #Return SQL Select statement with conditions to find discontinuing items on the website that are out of stock
    def _sql_select_items_to_remove(website_items, valid_items, valid_items_stock_avail):
        return f'''
          SELECT {website_items}.id, {website_items}.name, {valid_items_stock_avail}.`In Stock?`, {valid_items_stock_avail}.`Regular price`
          FROM website_items, valid_items, valid_items_stock_avail
            WHERE {website_items}.name = {valid_items}.product_id
              AND {website_items}.name = {valid_items_stock_avail}.product_id
              AND {valid_items_stock_avail}.`In Stock?`= 0
              AND {valid_items}.is_continuing = 0
        '''

    #Return SQL Select statement with conditions to find items that need to be updated due to out-dated price or inStock status
    def _sql_select_items_to_update(website_items, valid_items_stock_avail):
        return f'''
        SELECT {website_items}.id, {website_items}.name, {valid_items_stock_avail}.`In Stock?`, {valid_items_stock_avail}.`Regular price`
        FROM website_items, valid_items_stock_avail
        WHERE {website_items}.name = {valid_items_stock_avail}.product_id
          AND ({website_items}.in_stock != {valid_items_stock_avail}.`In Stock?`
            OR {website_items}.reg_price != {valid_items_stock_avail}.`Regular price`)
        '''

    #Returns SQL Select statement with conditions to find items to add onto website
    def _sql_select_items_to_add(website_items, valid_items, valid_items_stock_avail):
        return f'''
          SELECT {valid_items_stock_avail}.product_id, {valid_items_stock_avail}.`In Stock?`, {valid_items_stock_avail}.`Regular price`, {valid_items}.is_continuing
          FROM {valid_items_stock_avail}
          LEFT JOIN {valid_items} ON {valid_items_stock_avail}.product_id = {valid_items}.product_id
          WHERE (
            NOT EXISTS (
              SELECT name FROM {website_items}
                WHERE {website_items}.name = {valid_items_stock_avail}.product_id
              )
            )
            AND ({valid_items}.is_continuing = 1
            OR ({valid_items_stock_avail}.`Regular price` > 30
              AND {valid_items_stock_avail}.`In Stock?` = 1
            )
          )
          ORDER BY is_continuing, `In Stock?` DESC

        '''

    #Returns SQL Select statement with conditions to find items on website but not on accuterm
    def _sql_select_items_on_website_but_not_in_accuterm(website_items, valid_items_stock_avail):
        return f'''
          SELECT {website_items}.id, {website_items}.name, {valid_items_stock_avail}.`In Stock?`, {valid_items_stock_avail}.`Regular price`
          FROM website_items
          LEFT JOIN {valid_items_stock_avail} ON {website_items}.name = {valid_items_stock_avail}.product_id
          WHERE {valid_items_stock_avail}.product_id IS NULL
        '''

    #===================  INITIALIZE VARIABLES ========================================
    query_dict = {
      'sql_select_items_to_remove' : {
        'sql' : _sql_select_items_to_remove('website_items', 'valid_items', 'valid_items_stock_avail'),
        'description' : 'Automatically Removes Out of Stock Discontinued Items on Website'
      },
      'sql_select_items_to_update' : {
        'sql' : _sql_select_items_to_update('website_items', 'valid_items_stock_avail'),
        'description' : 'Automatically Updates Items on Website'
      },
      'sql_select_items_to_add' : {
        'sql' : _sql_select_items_to_add('website_items', 'valid_items', 'valid_items_stock_avail'),
        'description' : 'Displays Continued Items or In-stock Discontinued Items that are not on the Website'
      },
      'sql_select_items_on_website_but_not_in_accuterm' : {
        'sql' : _sql_select_items_on_website_but_not_in_accuterm('website_items', 'valid_items_stock_avail'),
        'description' : 'Displays  Items on Website but not in Accuterm'
      }
    }

    #===================  DEBUGGING FUNCTIONS ========================================

    #Creates '`query`.csv' file from passed in query as a list
    def create_csv_query(self, query_function):

        print(f'Creating Query: `{query_function}`.csv')

        with open(f"G:/Kevin/website info/Zach's Toolbox/Query_{query_function}.csv", "w", newline='') as csvfile:
          spamwriter = csv.writer(csvfile, delimiter=',',quotechar='|', quoting=csv.QUOTE_MINIMAL)
          self.cursor.execute('{};'.format(SQL.query_dict[query_function]['sql']))
          for row in self.cursor.fetchall():
            spamwriter.writerow([cell for cell in row])

    #Creates '`table_name`.csv' file from passed in table_name
    def create_csv_table(self, table_name):
        with open(f"G:\\Kevin\\website info\\Zach's Toolbox\\{table_name}.csv", "w", newline='') as csvfile:
          spamwriter = csv.writer(csvfile, delimiter=',',quotechar='|', quoting=csv.QUOTE_MINIMAL)
          self.cursor.execute(f'SELECT * FROM {table_name};')
          for row in self.cursor.fetchall():
            spamwriter.writerow([cell for cell in row])

    #===================  MAIN FUNCTIONS  =========================================

    #Gets Names/IDs and inserts into Table: 'website_items'
    def create_table_woocommerce_items(self, wcapi, per_page = 100):
          
        #Create new 'website_items' table
        try:
          print(f"Creating `website_items` Table")
          self._create_table('website_items')
        except sqlite3.Error as error:  self._print_error(error)

        total_items = 0

        response = wcapi.get('products')
        total_pages = int((int(response.headers['X-WP-TotalPages']) * 10) / per_page) + 1
        for i in range(1, total_pages + 1):
          print("{}%...".format(int((i * 100)/total_pages)))
          responseGET = wcapi.get('products?per_page={}&page='.format(per_page) + str(i)).json()
          for name in responseGET:
            total_items += 1
            self._insert_table('website_items', (name.get("id"), name.get("name"), int(name.get("stock_status") == 'instock'), name.get("regular_price")))

        self.cursor.execute('SELECT * FROM website_items;')
        results = self.cursor.fetchall()
        
        if total_items > len(results): 
          print("Error: Missing items due to fetched Duplicates from WooREST API")
          print(f"Missing {total_items - len(results)} items")

          #Automatically attempt to recover missing items once
          if per_page > 98:
            print("Attempting to Recover Missing Items...")
            self.create_table_woocommerce_items(wcapi, per_page - 2)

          #Provide User the option to proceed without missing item if already attempted to recover
          # else:
              

        else: print(f"Total Website Items: {total_items}")
      
        self.conn.commit()
      
    #Get Accuterm's Product Info from Exported Excel Workbook
    def create_tables_accuterm_items(self, file):

      #Dictionary of Column Field Names : Column Index
      col_dict = {
        'Product Description' : 4,
        'Description #2' : 5,
        'Description #3' : 6,
        'Reg Price' : 7,
        'P/U Price' : 8,
        'Avail' : 11,
      }

      #raw_items table serves as a placeholder for extracted products in excel sheet(s)
      self._create_table('raw_items')

      #Load worksheet object from file path
      wb_obj = load_workbook(filename=file)
      for sheet in wb_obj.sheetnames:
        wsheet = wb_obj[sheet]

        #Copy products from each worksheet into raw_items table
        for key, *values in wsheet.iter_rows():

          #Doesn't accept empty rolls and null P/U field value
          if key.value != ' ' and values[6].value != None:
            
            data_row = [v.value for v in values if v.column in col_dict.values() and v.row > 2]
            data_row = ['' if data == None else data for data in data_row]

            data_row.insert(0, key.value)

            #Insert True in field:'is_continuing' if item is continuing and false otherwise
            if wb_obj.sheetnames.index(sheet) <= 1: data_row.append(True)
            else:                                   data_row.append(False)

            self._insert_table('raw_items', tuple(data_row))

      #Copy/filter raw_items into valid_items table once reaches last sheet
      if wb_obj.sheetnames.index(sheet) == (len(wb_obj.sheetnames) - 1): 
        
        self.cursor.execute('CREATE TABLE valid_items AS {};'.format(self._sql_get_valid_items('raw_items')))

        #Truncates raw_item table
        self.cursor.execute('DELETE FROM raw_items;')

        #Insert into valid_items_stock_avail table with each items inStock status and price from valid_items table
        self.cursor.execute('CREATE TABLE valid_items_stock_avail AS {};'.format(self._sql_get_stock_price('valid_items')))

      self.conn.commit()

    #Return list of items from `query`
    def get_query_items(self, query):
      #Select ID's of discontinued items that are no longer in stock or have prices < $2 
      self.cursor.execute('{};'.format(SQL.query_dict[query]['sql']))
      return self.cursor.fetchall()

class MainWindow(QDialog):
    def __init__(self):
        super(MainWindow, self).__init__()
        loadUi("MainWindow.ui", self)
        self.button_file.clicked.connect(self.browseFiles)
    
    def browseFiles(self):
        fname=QFileDialog.getOpenFileName(self, 'Open file', filter="*.xlsx", directory="G:/")
        self.text_file.setText(fname[0])
        self.goto_LoadingScreen()

    def goto_LoadingScreen(self):
        loadingscreen=LoadingScreen(self.text_file.text())
        widget.addWidget(loadingscreen)
        widget.setCurrentIndex(widget.currentIndex()+1)

class LoadingScreen(QDialog):
    def __init__(self, filename):
        super(LoadingScreen, self).__init__()
        loadUi("LoadingScreen.ui", self)
        self.threadpool = QThreadPool()

        # Pass the function to execute
        self.worker = Worker(self.fetchItems, filename)
        self.worker.signals.progress.connect(self.progressValue)

        # Execute
        self.threadpool.start(self.worker)
    
    def progressValue(self, value):
        self.progressBar.setValue(value)
        if value == 100: self.goto_InfoScreen()

    def fetchItems(self, filename, progress_callback):
        try:
          #Initiate WooRest API class
          progress_callback.emit(0)
          self.text_file.setText('Initializing Woorest API Class')
          wcapi = API(
          url="url",
          consumer_key="consumer_key",
          consumer_secret="consumer_secret",
          timeout=50
          )
          progress_callback.emit(2)

          #Initialize SQL Class
          self.text_file.setText('Initializing SQLite Database')
          sql = SQL("BestQualityDatabase", drop_tables=True)

          #Randomize progress bar LMAO, didn't have time to figure out threadpools and call_back signals from the SQL class
          progress_callback.emit(randint(3, 7))

          #Create/Populate SQL table website items on woocommerce
          self.text_file.setText('Fetching Woocommerce Items\n(its not broken, itll take a while)')
          sql.create_table_woocommerce_items(wcapi)
          progress_callback.emit(randint(84, 96))

          #Create/Populate SQL table of Accuterm Items
          self.text_file.setText('Fetching Accuterm Items')
          sql.create_tables_accuterm_items(filename)
          progress_callback.emit(randint(96, 99))

          self.text_file.setText('Creating CSV Files from tables/queries')
          sql.create_csv_table('website_items')
          sql.create_csv_table('valid_items')
          sql.create_csv_table('valid_items_stock_avail')

          sql.create_csv_query('sql_select_items_to_remove')
          sql.create_csv_query('sql_select_items_to_update')
          sql.create_csv_query('sql_select_items_to_add')
          sql.create_csv_query('sql_select_items_on_website_but_not_in_accuterm')
          progress_callback.emit(100)

        except sqlite3.Error as err:
          print("Something went wrong: {}".format(err))

    def goto_InfoScreen(self):
        infoscreen=InfoScreen()
        widget.addWidget(infoscreen)
        widget.setCurrentIndex(widget.currentIndex()+1)
        widget.setFixedHeight(709)
        widget.setFixedWidth(569)

class InfoScreen(QDialog):
    def __init__(self):
        super(InfoScreen, self).__init__()
        loadUi("InfoScreen.ui", self)
        self.sql = SQL("BestQualityDatabase", drop_tables=False)
        self._init_tables()
        self.button_execute.clicked.connect(self.update_website)
        self.progressBar.hide()
        self.label_progress.hide()

    #Creates/populate tables from query dictionary
    def _init_tables(self):
      for query_key in self.sql.query_dict.keys():
        self.table = QTableWidget()
        
        self.sql.cursor.execute('{};'.format(self.sql.query_dict[query_key]['sql']))
        print(self.sql.cursor.description)

        #Add query tab with filtered title
        self.query_key = self.tabWidget.addTab(self.table, query_key.replace('_', ' ').title()[11:])

        #Add query tooltip
        self.tabWidget.setTabToolTip(self.tabWidget.count() - 1, self.sql.query_dict[query_key]['description'])

        if len(self.sql.get_query_items(query_key)) > 0:
              
          #Set row/column count
          self.table.setColumnCount(len(self.sql.get_query_items(query_key)[0]))
          self.table.setRowCount(len(self.sql.get_query_items(query_key)))

          #Set header label from query's column name
          self.table.setHorizontalHeaderLabels(list(map(lambda x: x[0], self.sql.cursor.execute('{}'.format(self.sql.query_dict[query_key]['sql'])).description)))

          #Populate table from query
          for row in self.sql.get_query_items(query_key):
            for cell in row:
              self.table.setItem(self.sql.get_query_items(query_key).index(row), row.index(cell), QTableWidgetItem(str(cell)))
        
      #Delete placeholder tab
      self.tabWidget.removeTab(0)
    
    #Make changes to website based on the status of each checkbox
    def update_website(self):
      self.button_execute.setEnabled(False)
      #Show progress widgets
      self.progressBar.show()
      self.label_progress.show()
      self.progressBar.setMaximum(len(self.sql.get_query_items('sql_select_items_to_remove')) + len(self.sql.get_query_items('sql_select_items_to_update')))

      wcapi = API(
          url="url",
          consumer_key="consumer_key",
          consumer_secret="consumer_secret",
          timeout=50
          )

      if self.checkbox_delete.isChecked(): 

        self.label_progress.setText('Deleting Items')

        #Delete discontinued out of stock items off website
        for row in self.sql.get_query_items('sql_select_items_to_remove'):
            response = wcapi.delete(f'products/{row[0]}')
            self.progressBar.setValue(self.progressBar.value() + 1)

      if self.checkbox_update.isChecked(): 

        stock_status_dict = {0 : 'outofstock', 1 : 'instock'}

        self.label_progress.setText('Updating Items')

        #Update Items on website
        for row in self.sql.get_query_items('sql_select_items_to_update'):
            data = {
              "stock_status": stock_status_dict.get(int(row[2])),
              "regular_price": str(row[3])
            }
            print(wcapi.put(f'products/{row[0]}', data).json)
            self.progressBar.setValue(self.progressBar.value() + 1)

      self.progressBar.setValue(self.progressBar.maximum())
      self.label_progress.setText('Done')
#=================================================================================

app=QApplication(sys.argv)
mainwindow=MainWindow()
widget=QtWidgets.QStackedWidget()
widget.addWidget(mainwindow)
widget.setWindowTitle("Zach's Magic Box")
widget.setFixedHeight(750)
widget.setFixedWidth(700)
widget.show()
sys.exit(app.exec())